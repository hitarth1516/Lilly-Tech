import json
import csv
import os
import pandas as pd
import re
import sys
import logging
from openpyxl import Workbook
from datetime import datetime

# Configure logging
timestmp = datetime.now().strftime("%Y%m%d_%H%M%S")
log_filename = f'./validate_files_{timestmp}.log'
logging.basicConfig(filename=log_filename, filemode='w', level=logging.INFO)

# --- Start of integrated trim_spaces.py functionality ---

def trim_csv_spaces(file_path, sep, quoting):
    """
    Reads a CSV file, trims spaces from all string columns,
    and overwrites the original file.
    Accepts delimiter (sep) and quoting style.
    """
    try:
        if not os.path.exists(file_path):
            logging.error(f"Error: File not found for trimming: {file_path}")
            print(f"Error: File not found for trimming: {file_path}")
            # Do not exit here, let the main validation handle file existence
            return False

        # Read the CSV file using provided delimiter and quoting
        df = pd.read_csv(file_path, sep=sep, quoting=quoting)
        logging.info(f"Successfully read CSV for trimming: {file_path}")

        # Trim spaces from all string columns
        for col in df.select_dtypes(include=['object']).columns:
            # Ensure the column is treated as string before stripping
            df[col] = df[col].astype(str).str.strip()
        logging.info("Trimmed spaces from string columns.")

        # Overwrite the original CSV file with trimmed data
        df.to_csv(file_path, sep=sep, index=False, quoting=quoting)
        logging.info(f"Successfully wrote trimmed data back to: {file_path}")
        print(f"Successfully trimmed spaces in {file_path}")
        return True

    except Exception as e:
        logging.error(f"Failed to trim spaces in {file_path}: {e}")
        print(f"Failed to trim spaces in {file_path}: {e}")
        return False

# --- End of integrated trim_spaces.py functionality ---


def is_valid_csv(file_path, sep, quoting):
    """Validate CSV file for format and invalid characters."""
    try:
        with open(file_path, 'r', encoding='utf-8') as f:
            content = f.read()

        df = pd.read_csv(file_path, sep=sep, quoting=quoting)
        if df.empty:
            print(f"Warning: Empty CSV file: {file_path}")
            logging.warning(f"Empty CSV file: {file_path}")
            return False
        else:
            print(f"Valid CSV: {file_path}")
            logging.info(f"Valid CSV: {file_path}")
            return True
    except pd.errors.ParserError as e:
        print(f"Error: Invalid CSV format in {file_path}: {e}")
        logging.error(f"Invalid CSV format in {file_path}: {e}")
        return False
    except UnicodeDecodeError as e:
        print(f"Error: Encoding issue in {file_path}: {e}")
        logging.error(f"Encoding issue in {file_path}: {e}")
        return False

def check_duplicates(metric_config_df, master_config_df):
    """Check for duplicates in master and metric config files."""
    status1, status2 = 'SUCCESS', 'SUCCESS'
    master_config_active_query_df = master_config_df[master_config_df['TYPE'] == 'QUERY']
    metric_config_active_entry_df = metric_config_df[['PARAMETER_VALUE', 'CONFIG_TYPE', 'DESCRIPTION']]

    filtered_df = master_config_active_query_df[
        ~master_config_active_query_df['Identifier_Value1'].isin(['NPP_OBU_MASTER_DERIVED', 'NPP_OBU_MASTER_PASSTHROUGH'])
    ]
    duplicates_df1 = pd.DataFrame() if not filtered_df['Identifier_Value1'].duplicated().any() else \
        pd.DataFrame(filtered_df[filtered_df['Identifier_Value1'].duplicated(keep=False)]['Identifier_Value1'].unique(), columns=['Duplicate IdentifierValue1 Values'])
    if not duplicates_df1.empty:
        status1 = 'FAILED'

    filtered_df_obu = master_config_active_query_df[
        master_config_active_query_df['Identifier_Value1'].isin(['NPP_OBU_MASTER_DERIVED', 'NPP_OBU_MASTER_PASSTHROUGH'])
    ]
    if filtered_df_obu.duplicated(subset=['Identifier_Value1', 'Identifier_Value2']).any():
        duplicates_obu = filtered_df_obu[filtered_df_obu.duplicated(subset=['Identifier_Value1', 'Identifier_Value2'], keep=False)][['Identifier_Value1', 'Identifier_Value2']].drop_duplicates()
        duplicates_df1 = pd.concat([duplicates_df1, duplicates_obu['Identifier_Value1']], ignore_index=True)
        status1 = 'FAILED'

    duplicates_df2 = pd.DataFrame() if not metric_config_active_entry_df.duplicated().any() else \
        metric_config_active_entry_df[metric_config_active_entry_df.duplicated(keep=False)].copy()
    if not duplicates_df2.empty:
        status2 = 'FAILED'

    return {
        'MasterConfig_DuplicateCheck': (status1, duplicates_df1),
        'MetricConfig_DuplicateCheck': (status2, duplicates_df2)
    }

def validate_master_files(master_config_df):
    """Validate presence of expected master files."""
    status = 'SUCCESS'
    master_file_list = ['customer', 'product', 'rep_alignment', 'intermediate_output', 'npp_output', 'pp_output', 'seg_output']
    active_files = master_config_df[(master_config_df['TYPE'] == 'INPUT_FILE') | (master_config_df['TYPE'] == 'OUTPUT_FILE')]['Identifier_Name1'].dropna().unique()
    missing_files = [f for f in master_file_list if f.lower() not in [file.lower() for file in active_files]]
    missing_files_df = pd.DataFrame(missing_files, columns=['Master Input Missing Files']) if missing_files else pd.DataFrame()
    if missing_files:
        status = 'FAILED'
    return {'MissingFilesCheck': (status, missing_files_df)}

def extract_table_names(sql_query):
    """Extract table names from SQL queries."""
    # Ensure sql_query is treated as a string, handling potential non-string types (like floats for NaN)
    table_name_pattern = re.compile(r'\bFROM\s+(\w+)|\bJOIN\s+(\w+)', re.IGNORECASE)
    matches = table_name_pattern.findall(str(sql_query)) # Convert to string here
    return list(set(name for match in matches for name in match if name))

def validate_table_names(master_config_df):
    """Validate tables used in SQL queries."""
    status = 'SUCCESS'
    active_df = master_config_df.copy()
    # Ensure 'Identifier_Value2' is treated as string before applying extract_table_names
    active_df['Extracted_Tables'] = active_df['Identifier_Value2'].astype(str).apply(lambda x: extract_table_names(x) if pd.notnull(x) else [])
    extracted_tables = [t.lower() for sublist in active_df['Extracted_Tables'] for t in sublist if t.lower() != 'temp']
    unique_values = [v.lower() for v in master_config_df['File_Name'].dropna().unique()]
    missing_tables = [t for t in extracted_tables if t not in unique_values]
    missing_tables_df = pd.DataFrame(missing_tables, columns=['SQL INPUT MISSING TABLES']) if missing_tables else pd.DataFrame()
    if missing_tables:
        status = 'FAILED'
    return {'MissingSQLTablesCheck': (status, missing_tables_df)}

def print_inactive_files(master_config_df):
    """List inactive files in master config."""
    status = 'SUCCESS' if len(master_config_df[master_config_df["Status"] == 'InActive']['File_Name'].dropna().unique()) == 0 else 'USER_EVALUATE'
    inactive_files_df = pd.DataFrame(master_config_df[master_config_df["Status"] == 'InActive']['File_Name'].dropna().unique(), columns=['INACTIVE INPUT FILE LIST'])
    return {'InactiveInputFiles-MasterConfig': (status, inactive_files_df)}

def print_inactive_metrics(master_config_df):
    """List inactive metrics in master config."""
    status = 'SUCCESS' if len(master_config_df[master_config_df["Status"] == 'InActive']['Identifier_Name2'].dropna().unique()) == 0 else 'USER_EVALUATE'
    inactive_metrics_df = pd.DataFrame(master_config_df[master_config_df["Status"] == 'InActive']['Identifier_Name2'].dropna().unique(), columns=['INACTIVE INPUT FILE LIST'])
    return {'InactiveScripts-MasterConfig': (status, inactive_metrics_df)}

def validate_sql_template_strings(master_config_df):
    """Validate template strings against dynamic statements."""
    status = 'SUCCESS'
    failure_df = pd.DataFrame(columns=['IdentifierValue', 'Number of Template Strings', 'Number of Dynamic Statements / Trigger Names'])
    for index, row in master_config_df.iterrows():
        if row['Identifier_Name1'] in ('NPP', 'PP', 'SEG'):
            # Ensure 'Identifier_Value2' and 'DynamicStatements' are strings
            identifier_value2_str = str(row['Identifier_Value2']) if pd.notnull(row['Identifier_Value2']) else ''
            dynamic_statements_str = str(row['DynamicStatements']) if pd.notnull(row['DynamicStatements']) else ''

            template_pattern = re.compile(r"#template_string_\d+")
            unique_templates = set(template_pattern.findall(identifier_value2_str))
            num_templates = len(unique_templates)
            num_statements = len([s for s in dynamic_statements_str.split('||') if s.strip()])
            if num_templates != num_statements:
                status = 'FAILED'
                failure_df = pd.concat([failure_df, pd.DataFrame([{
                    'IdentifierValue': row['Identifier_Value1'],
                    'Number of Template Strings': num_templates,
                    'Number of Dynamic Statements / Trigger Names': num_statements
                }])], ignore_index=True)
    return {'TemplateString-DynamicStatement': (status, failure_df)}

def validate_quotes_in_sql(master_config_df):
    """Validate double quotes in SQL queries."""
    status = 'SUCCESS'
    failure_df = pd.DataFrame(columns=['IdentifierName', 'IdentifierValue'])
    for index, row in master_config_df.iterrows():
        if row['Identifier_Name1'] in ('NPP', 'PP', 'SEG'):
            # Ensure 'Identifier_Value2' is a string
            identifier_value2_str = str(row['Identifier_Value2']) if pd.notnull(row['Identifier_Value2']) else ''
            if identifier_value2_str and (identifier_value2_str.startswith('"') or identifier_value2_str.endswith('"')):
                status = 'FAILED'
                failure_df = pd.concat([failure_df, pd.DataFrame([{
                    'IdentifierName': row['Identifier_Name2'],
                    'IdentifierValue': row['Identifier_Value1']
                }])], ignore_index=True)
    return {'DoubleQuoteSQLScripts': (status, failure_df)}

def validate_quotes_in_metric_config(metric_config_df):
    """Validate double quotes in metric config values."""
    status = 'SUCCESS'
    failure_df = pd.DataFrame(columns=['ParameterValue', 'ConfigType'])
    for index, row in metric_config_df.iterrows():
        if row['PARAMETER_TYPE'] == 'DYNAMIC_METRIC_ID':
            # Ensure 'CONFIG_VALUE' is a string
            config_value_str = str(row['CONFIG_VALUE']) if pd.notnull(row['CONFIG_VALUE']) else ''
            if config_value_str and (config_value_str.startswith('"') or config_value_str.endswith('"')):
                status = 'FAILED'
                failure_df = pd.concat([failure_df, pd.DataFrame([{
                    'ParameterValue': row['PARAMETER_VALUE'],
                    'ConfigType': row['CONFIG_TYPE']
                }])], ignore_index=True)
    return {'DoubleQuoteMetricNames': (status, failure_df)}

def validate_active_metrics(metric_config_df, env_data):
    """Validate active metrics against environment data (simplified)."""
    status = 'SUCCESS'
    # Assuming env_data has an 'active_metrics' key; adjust if different
    env_metrics = env_data.get('active_metrics', [])
    active_metrics = metric_config_df[metric_config_df['ACTIVE'] == 'Y']['DESCRIPTION'].dropna().unique()
    discrepancies = [m for m in active_metrics if m not in env_metrics]
    discrepancy_df = pd.DataFrame(discrepancies, columns=['Metrics-InMetricAsset-NotinENV']) if discrepancies else pd.DataFrame()
    if discrepancies:
        status = 'FAILED'
    return {'Metrics-InMetricAsset-NotinENV': (status, discrepancy_df)}

def print_inactive_metrics_list(metric_config_df):
    """List inactive metrics in metric config."""
    status = 'SUCCESS' if len(metric_config_df[metric_config_df["ACTIVE"] == 'N']['DESCRIPTION'].dropna().unique()) == 0 else 'USER_EVALUATE'
    inactive_metrics_df = pd.DataFrame(metric_config_df[metric_config_df["ACTIVE"] == 'N']['DESCRIPTION'].dropna().unique(), columns=['INACTIVE METRICS LIST'])
    return {'InactiveMetrics-MetricConfig': (status, inactive_metrics_df)}

def validate_activeinsights_metrics(master_config_df, metric_config_df, env_data):
    """Validate metrics in active insights (simplified)."""
    status1, status2 = 'SUCCESS', 'SUCCESS'
    active_master_df = pd.DataFrame()
    for index, row in master_config_df.iterrows():
        # Ensure 'DynamicStatements' is a string before splitting
        dynamic_statements_str = str(row['DynamicStatements']) if pd.notnull(row['DynamicStatements']) else ''
        if row['Identifier_Name1'] in ('NPP', 'PP', 'SEG') and dynamic_statements_str: # Check if string is not empty after conversion
            statements = dynamic_statements_str.split('||')
            names = [s.split('#')[0] for s in statements]
            active_master_df = pd.concat([active_master_df, pd.DataFrame({
                'Identifier_Value1': row['Identifier_Value1'],
                'NameBeforeHash': [n + '#' for n in names]
            })], ignore_index=True)

    merged_df = pd.merge(active_master_df, metric_config_df, left_on=['Identifier_Value1', 'NameBeforeHash'],
                        right_on=['CONFIG_TYPE', 'PARAMETER_VALUE'], how='inner')
    metric_asset_active = merged_df[merged_df['ACTIVE'] == 'Y']['DESCRIPTION']
    env_active = env_data.get('active_metrics', [])
    non_matching = [m for m in metric_asset_active if m not in env_active]
    non_matching_df = pd.DataFrame(non_matching, columns=['Metrics-InOC-NotActiveMetrics']) if non_matching else pd.DataFrame()
    if non_matching:
        status1 = 'FAILED'

    result_df = pd.DataFrame(columns=['suggestionName', 'colName'])  # Placeholder; adjust with actual data
    if not result_df.empty:
        status2 = 'FAILED'

    return {
        'Metrics-InOC-NotActiveMetrics': (status1, non_matching_df),
        'Metrics-InOC-NotMetricAsset': (status2, result_df)
    }

def split_dynamic_statements_templateid(row):
    """Split dynamic statements to extract template IDs."""
    # Ensure 'DynamicStatements' is a string before splitting
    dynamic_statements_str = str(row['DynamicStatements']) if pd.notnull(row['DynamicStatements']) else ''
    split_df = pd.DataFrame(columns=['IdentifierValue', 'TemplateString'])
    if not dynamic_statements_str: # Check if string is empty after conversion
        return split_df
    statements = dynamic_statements_str.split('||')
    for part in statements:
        if '#template_id' in part:
            metric_name_match = re.search(r'(\w+)#', part)
            if metric_name_match:
                template_string = metric_name_match.group(1)
                split_df = pd.concat([split_df, pd.DataFrame([{
                    'IdentifierValue': row['Identifier_Value1'],
                    'TemplateString': template_string
                }])], ignore_index=True)
    return split_df

def validate_dynamic_statements_vaetemplateids(master_config_df, metric_config_df):
    """Validate VAE template IDs in dynamic statements."""
    status = 'SUCCESS'
    failure_df = pd.DataFrame(columns=['IdentifierValue', 'TemplateString'])
    master_config_active_query_df = master_config_df[master_config_df['TYPE'] == 'QUERY']
    master_config_validation_df = master_config_active_query_df[['Identifier_Value1', 'DynamicStatements']]
    metric_config_validation_df = metric_config_df[metric_config_df['PARAMETER_TYPE'] == 'DYNAMIC_METRIC_ID'][['PARAMETER_VALUE', 'CONFIG_TYPE', 'CONFIG_VALUE']].drop_duplicates()

    expanded_rows = []
    for _, row in master_config_validation_df.iterrows():
        split_df = split_dynamic_statements_templateid(row)
        expanded_rows.append(split_df)
    expanded_df = pd.concat(expanded_rows, ignore_index=True) if expanded_rows else pd.DataFrame(columns=['IdentifierValue', 'TemplateString'])
    master_config_expanded_df = expanded_df[['IdentifierValue', 'TemplateString']].drop_duplicates()

    for index, row in master_config_expanded_df.iterrows():
        count = 0
        for index1, row1 in metric_config_validation_df.iterrows():
            if (row['IdentifierValue'] == row1['CONFIG_TYPE']) and (row['TemplateString'] + '#' == row1['PARAMETER_VALUE']):
                # Ensure 'CONFIG_VALUE' is a string
                config_value_str = str(row1['CONFIG_VALUE']) if pd.notnull(row1['CONFIG_VALUE']) else ''
                if '"template_id"' in config_value_str:
                    count += 1
        if count == 0:
            status = 'FAILED'
            failure_df = pd.concat([failure_df, pd.DataFrame([{
                'IdentifierValue': row['Identifier_Value1'],
                'TemplateString': row['TemplateString'] + '#'
            }])], ignore_index=True)

    return {'MASTER VS METRIC TEMPLATEID': (status, failure_df.drop_duplicates())}

def flag_double_quote(master_config_df, metric_config_df):
    """Flag rows with double or single quotes at start/end."""
    status1, status2 = 'SUCCESS', 'SUCCESS'
    failure_df1 = pd.DataFrame(columns=['ID Number'])
    failure_df2 = pd.DataFrame(columns=['ROW DESCRIPTION NAME'])

    master_check_df = master_config_df.copy() # Operate on a copy to avoid SettingWithCopyWarning
    metric_check_df = metric_config_df.copy() # Operate on a copy to avoid SettingWithCopyWarning

    # Apply string conversion before checking for quotes
    mask1 = master_check_df.apply(lambda col: col.astype(str).str.contains('""') | col.astype(str).str.startswith('"') | col.astype(str).str.endswith('"'))
    mask2 = metric_check_df.apply(lambda col: col.astype(str).str.contains('""') | col.astype(str).str.startswith('"') | col.astype(str).str.endswith('"'))

    master_check_df['Flag'] = mask1.any(axis=1)
    metric_check_df['Flag'] = mask2.any(axis=1)

    for index, row in master_check_df.iterrows():
        if row['Flag']:
            status1 = 'FAILED'
            failure_df1 = pd.concat([failure_df1, pd.DataFrame([{'ID Number': row.get('ID', 'N/A')}])], ignore_index=True)
    for index, row in metric_check_df.iterrows():
        if row['Flag']:
            status2 = 'FAILED'
            failure_df2 = pd.concat([failure_df2, pd.DataFrame([{'ROW DESCRIPTION NAME': row.get('DESCRIPTION', 'N/A')}])], ignore_index=True)

    return {
        'MASTER CONFIG DOUBLEQUOTES ROW': (status1, failure_df1),
        'METRIC CONFIG DOUBLEQUOTES ROW': (status2, failure_df2)
    }

def generate_report(results, output_path):
    """Generate an Excel report with validation results."""
    wb = Workbook()
    # Remove the default sheet created by Workbook()
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])

    # Add a summary sheet for initial file validation
    ws_summary = wb.create_sheet(title="File Validation Summary")
    ws_summary.append(["File", "Status", "Details"])
    for file_path, (status, details) in results.get('initial_validation', {}).items():
        ws_summary.append([os.path.basename(file_path), status, details])

    # Add sheets for each additional validation check
    for sheet_name, (status, df) in results.items():
        if sheet_name == 'initial_validation': # Skip the initial validation as it's handled in summary
            continue

        ws = wb.create_sheet(title=sheet_name)
        ws.append([f"{sheet_name} Status", status]) # Add overall status for the check

        if not df.empty:
            # Append column headers
            ws.append(df.columns.tolist())
            # Append data rows
            for r_idx, row in df.iterrows():
                ws.append(row.tolist())
        else:
            ws.append(["No issues found for this check."])

    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    wb.save(output_path)
    logging.info(f"Validation report saved to {output_path}")
    print(f"Validation report saved to {output_path}")

def main():
    """Main function to validate specific CSV files and perform additional checks."""
    try:
        # Get file paths from environment variables
        metric_config_path = os.environ.get('METRIC_CONFIG_PATH', 'INPUT/03_METADATA_FILES/METRIC_SEGMENT_ASSET/CONFIG_FILE/METRIC_CONFIG/Metric_Config_File.csv')
        master_config_path = os.environ.get('MASTER_CONFIG_PATH', 'INPUT/03_METADATA_FILES/METRIC_SEGMENT_ASSET/CONFIG_FILE/MASTER_CONFIG/Master_Config_File.csv')
        output_path = os.environ.get('OUTPUT_PATH', 'INPUT/DQM_REPORT/dqm_report.xlsx')

        # Define parameters for each file type
        file_processing_params = {
            metric_config_path: {'sep': '|', 'quoting': csv.QUOTE_NONE},
            master_config_path: {'sep': ',', 'quoting': csv.QUOTE_MINIMAL}
        }

        results = {'initial_validation': {}}
        all_valid = True

        # --- Trim spaces before validation and reading ---
        print("Starting trim spaces for metric config file...")
        # Call trim_csv_spaces for metric_config_path using its specific parameters
        trim_success = trim_csv_spaces(metric_config_path,
                                       file_processing_params[metric_config_path]['sep'],
                                       file_processing_params[metric_config_path]['quoting'])
        if not trim_success:
            print(f"Warning: Skipping further validation due to failure in trimming {metric_config_path}.")
            logging.warning(f"Skipping further validation due to failure in trimming {metric_config_path}.")
            all_valid = False # Mark overall as invalid if trimming fails

        # --- Initial validation of CSV files ---
        # Only proceed with initial validation if trimming was successful for metric_config_path
        if all_valid:
            for file_path, params in file_processing_params.items():
                if os.path.exists(file_path):
                    status = is_valid_csv(file_path, params['sep'], params['quoting'])
                    details = "Validation passed" if status else "Validation failed"
                    results['initial_validation'][file_path] = ("SUCCESS" if status else "FAILED", details)
                    if not status:
                        all_valid = False
                else:
                    print(f"Error: File not found: {file_path}")
                    logging.error(f"File not found: {file_path}")
                    results['initial_validation'][file_path] = ("FAILED", "File not found")
                    all_valid = False

        # Read data for additional validations
        metric_config_df = pd.DataFrame()
        master_config_df = pd.DataFrame()

        # Only read dataframes if initial validation passed for them
        if results['initial_validation'].get(metric_config_path, ('FAILED', ''))[0] == 'SUCCESS':
            metric_config_df = pd.read_csv(metric_config_path,
                                           sep=file_processing_params[metric_config_path]['sep'],
                                           quoting=file_processing_params[metric_config_path]['quoting'])
        else:
            print(f"Skipping further validation: {metric_config_path} is invalid or not found.")
            logging.warning(f"Skipping further validation: {metric_config_path} is invalid or not found.")
            all_valid = False

        if results['initial_validation'].get(master_config_path, ('FAILED', ''))[0] == 'SUCCESS':
            master_config_df = pd.read_csv(master_config_path,
                                           sep=file_processing_params[master_config_path]['sep'],
                                           quoting=file_processing_params[master_config_path]['quoting'])
        else:
            print(f"Skipping further validation: {master_config_path} is invalid or not found.")
            logging.warning(f"Skipping further validation: {master_config_path} is invalid or not found.")
            all_valid = False

        env_data = {}  # Placeholder; adjust with actual env data source

        # Run additional validations only if both config dataframes are not empty
        if not metric_config_df.empty and not master_config_df.empty:
            results.update(check_duplicates(metric_config_df, master_config_df))
            results.update(validate_master_files(master_config_df))
            results.update(validate_table_names(master_config_df))
            results.update(print_inactive_files(master_config_df))
            results.update(print_inactive_metrics(master_config_df))
            results.update(validate_sql_template_strings(master_config_df))
            results.update(validate_quotes_in_sql(master_config_df))
            results.update(validate_quotes_in_metric_config(metric_config_df))
            results.update(validate_active_metrics(metric_config_df, env_data))
            results.update(print_inactive_metrics_list(metric_config_df))
            results.update(validate_activeinsights_metrics(master_config_df, metric_config_df, env_data))
            results.update(validate_dynamic_statements_vaetemplateids(master_config_df, metric_config_df))
            results.update(flag_double_quote(master_config_df, metric_config_df))
        else:
            print("Skipping additional validations due to missing or invalid configuration files.")
            logging.warning("Skipping additional validations due to missing or invalid configuration files.")


        # Generate report
        generate_report(results, output_path)

        if not all_valid:
            sys.exit(1)
        print("All validated files and checks are valid!")

    except Exception as e:
        logging.error(f"Execution failed: {e}")
        print(f"Execution failed: {e}")
        sys.exit(1)
    finally:
        logging.shutdown()

if __name__ == "__main__":
    main()

